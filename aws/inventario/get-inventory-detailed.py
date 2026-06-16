#!/usr/bin/env python3

"""Collect AWS resource inventory enriched with service family and size metadata."""

import argparse
import json
import math
import os
import sys
from collections import defaultdict
from pathlib import Path

try:
    import boto3
    from botocore.exceptions import BotoCoreError, ClientError
except ImportError as exc:
    sys.stderr.write(f"boto3 is required to run this script: {exc}\n")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    sys.stderr.write(f"openpyxl is required to run this script: {exc}\n")
    sys.exit(1)

DEFAULT_OUTPUT = os.environ.get(
    "OUTPUT_DIR", os.path.expanduser("~/Download/inventario/aws")
)


def build_parser():
    parser = argparse.ArgumentParser(
        description=(
            "Coleta inventário de recursos AWS, enriquecendo cada item com a família do serviço "
            "e informações de tamanho/capacidade quando disponíveis. Os relatórios são gerados "
            "em JSON e Excel, e podem ser consultados via subcomandos."
        )
    )

    subparsers = parser.add_subparsers(dest="command", required=True)

    collect_parser = subparsers.add_parser(
        "collect", help="Coleta o inventário e gera os relatórios enriquecidos."
    )
    collect_parser.add_argument(
        "--regions",
        nargs="*",
        help="Lista de regiões AWS. Padrão: todas as regiões suportadas pela API.",
    )
    collect_parser.add_argument(
        "--profile",
        help="Perfil de credenciais AWS a utilizar (opcional).",
    )
    collect_parser.add_argument(
        "--output-dir",
        default=DEFAULT_OUTPUT,
        help="Diretório de saída para os relatórios (padrão: ~/Download/inventario/aws).",
    )
    collect_parser.add_argument(
        "--page-size",
        type=int,
        default=100,
        help="Quantidade de recursos por requisição à API (máximo 100).",
    )

    query_parser = subparsers.add_parser(
        "query", help="Consulta relatórios existentes filtrando por serviço ou tipo."
    )
    query_parser.add_argument(
        "service",
        help="Serviço (ex.: ec2) ou tipo completo (ex.: ec2:instance) a filtrar.",
    )
    query_parser.add_argument(
        "--regions",
        nargs="*",
        help="Filtra resultados por região (opcional).",
    )
    query_parser.add_argument(
        "--output-dir",
        default=DEFAULT_OUTPUT,
        help="Diretório contendo os relatórios gerados (padrão: ~/Download/inventario/aws).",
    )

    return parser


def ensure_output_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def detect_regions(session: boto3.Session, cli_regions: list[str] | None) -> list[str]:
    if cli_regions:
        return cli_regions
    available = session.get_available_regions("resourcegroupstaggingapi")
    if not available:
        raise RuntimeError("Nenhuma região disponível para resourcegroupstaggingapi.")
    return available


def resource_type_from_arn(arn: str) -> str:
    if not arn or not arn.startswith("arn:"):
        return "unknown"
    parts = arn.split(":", 5)
    if len(parts) < 6:
        return "unknown"
    service = parts[2] or "unknown"
    resource = parts[5]
    primary = ""
    if service == "s3" and resource and "/" not in resource and ":" not in resource:
        primary = "bucket"
    else:
        for separator in ("/", ":"):
            if separator in resource:
                primary = resource.split(separator, 1)[0]
                break
        if not primary:
            primary = resource if resource else "resource"
    return f"{service}:{primary}"


def collect_region_inventory(session: boto3.Session, region: str, page_size: int):
    client = session.client("resourcegroupstaggingapi", region_name=region)
    grouped: dict[str, dict[str, dict]] = defaultdict(dict)
    duplicates = 0
    paginator = client.get_paginator("get_resources")

    for page in paginator.paginate(ResourcesPerPage=page_size):
        for mapping in page.get("ResourceTagMappingList", []):
            arn = mapping.get("ResourceARN")
            if not arn:
                continue
            resource_type = resource_type_from_arn(arn)
            tags = {}
            for tag in mapping.get("Tags", []):
                key = tag.get("Key")
                if key is None:
                    continue
                tags[key] = tag.get("Value", "")

            existing = grouped[resource_type].get(arn)
            if existing:
                duplicates += 1
                if tags:
                    existing_tags = existing.get("tags") or {}
                    existing_tags.update(tags)
                    existing["tags"] = existing_tags
                continue

            grouped[resource_type][arn] = {"arn": arn, "tags": tags}

    grouped_lists = {
        rtype: list(entries.values()) for rtype, entries in grouped.items()
    }
    total_unique = sum(len(entries) for entries in grouped_lists.values())
    return grouped_lists, total_unique, duplicates


def resource_display_name(arn: str, tags: dict | None) -> str:
    if tags:
        name = tags.get("Name") or tags.get("name")
        if name:
            return name
    if not arn:
        return ""
    resource_id = arn.rsplit("/", 1)[-1]
    resource_id = resource_id.rsplit(":", 1)[-1]
    return resource_id


def format_tags(tags: dict | None) -> str:
    if not tags:
        return ""
    return ";".join(
        f"{key}={'' if value is None else value}"
        for key, value in sorted(tags.items())
    )


def enrich_inventory(session: boto3.Session, region: str, inventory: dict[str, list[dict]]):
    for resource_type, entries in inventory.items():
        family = resource_type.split(":", 1)[0]
        for entry in entries:
            entry.setdefault("tags", {})
            entry.setdefault("details", {})
            entry.setdefault("size", "")
            entry["family"] = family
            entry.setdefault("resource_name", resource_display_name(entry.get("arn", ""), entry.get("tags")))

    enrich_ec2(session, region, inventory)
    enrich_rds(session, region, inventory)
    enrich_lambda(session, region, inventory)
    enrich_sqs(session, region, inventory)
    enrich_sns(session, region, inventory)


def enrich_ec2(session: boto3.Session, region: str, inventory: dict[str, list[dict]]):
    entries = inventory.get("ec2:instance")
    if not entries:
        return

    client = session.client("ec2", region_name=region)
    id_to_entry: dict[str, dict] = {}
    for entry in entries:
        arn = entry.get("arn", "")
        instance_id = arn.rsplit("/", 1)[-1]
        if instance_id:
            id_to_entry[instance_id] = entry

    if not id_to_entry:
        return

    instance_types: dict[str, dict] = {}

    ids = list(id_to_entry.keys())
    for i in range(0, len(ids), 100):
        chunk = ids[i : i + 100]
        try:
            paginator = client.get_paginator("describe_instances")
            for page in paginator.paginate(InstanceIds=chunk):
                for reservation in page.get("Reservations", []):
                    for instance in reservation.get("Instances", []):
                        entry = id_to_entry.get(instance.get("InstanceId", ""))
                        if not entry:
                            continue
                        instance_type = instance.get("InstanceType", "")
                        entry["details"].update(
                            {
                                "instance_type": instance_type,
                                "state": instance.get("State", {}).get("Name"),
                                "availability_zone": instance.get("Placement", {}).get("AvailabilityZone"),
                            }
                        )
                        entry["size"] = instance_type or entry["size"]
                        if instance_type and instance_type not in instance_types:
                            instance_types[instance_type] = {}
        except (ClientError, BotoCoreError) as exc:
            sys.stderr.write(f"[WARN] {region} EC2 describe_instances falhou: {exc}\n")
            break

    if not instance_types:
        return

    try:
        paginator = client.get_paginator("describe_instance_types")
        for page in paginator.paginate(InstanceTypes=list(instance_types.keys())):
            for info in page.get("InstanceTypes", []):
                instance_type = info.get("InstanceType")
                if not instance_type:
                    continue
                memory_mib = info.get("MemoryInfo", {}).get("SizeInMiB")
                vcpus = info.get("VCpuInfo", {}).get("DefaultVCpus")
                storage = info.get("InstanceStorageInfo", {}).get("TotalSizeInGB")
                instance_types[instance_type] = {
                    "memory_mib": memory_mib,
                    "memory_gb": round(memory_mib / 1024, 2) if memory_mib else None,
                    "vcpus": vcpus,
                    "instance_storage_gb": storage,
                }
    except (ClientError, BotoCoreError) as exc:
        sys.stderr.write(f"[WARN] {region} EC2 describe_instance_types falhou: {exc}\n")
        return

    for entry in entries:
        instance_type = entry.get("details", {}).get("instance_type")
        if not instance_type:
            continue
        specs = instance_types.get(instance_type)
        if not specs:
            continue
        entry["details"].update(specs)
        parts = [instance_type]
        if specs.get("memory_gb") is not None:
            parts.append(f"{specs['memory_gb']} GiB RAM")
        if specs.get("vcpus") is not None:
            parts.append(f"{specs['vcpus']} vCPU")
        entry["size"] = ", ".join(parts)


def enrich_rds(session: boto3.Session, region: str, inventory: dict[str, list[dict]]):
    entries = inventory.get("rds:db")
    if not entries:
        return

    client = session.client("rds", region_name=region)
    arn_map = {entry.get("arn", ""): entry for entry in entries if entry.get("arn")}
    if not arn_map:
        return

    try:
        paginator = client.get_paginator("describe_db_instances")
        for page in paginator.paginate():
            for db in page.get("DBInstances", []):
                arn = db.get("DBInstanceArn")
                entry = arn_map.get(arn)
                if not entry:
                    continue
                instance_class = db.get("DBInstanceClass", "")
                allocated = db.get("AllocatedStorage")
                max_allocated = db.get("MaxAllocatedStorage")
                entry["details"].update(
                    {
                        "db_instance_class": instance_class,
                        "allocated_storage_gib": allocated,
                        "max_allocated_storage_gib": max_allocated,
                        "engine": db.get("Engine"),
                        "multi_az": db.get("MultiAZ"),
                    }
                )
                size_parts = [instance_class] if instance_class else []
                if allocated:
                    size_parts.append(f"{allocated} GiB")
                if max_allocated and max_allocated != allocated:
                    size_parts.append(f"até {max_allocated} GiB")
                entry["size"] = " | ".join(size_parts)
    except (ClientError, BotoCoreError) as exc:
        sys.stderr.write(f"[WARN] {region} RDS describe_db_instances falhou: {exc}\n")


def enrich_lambda(session: boto3.Session, region: str, inventory: dict[str, list[dict]]):
    entries = inventory.get("lambda:function")
    if not entries:
        return

    client = session.client("lambda", region_name=region)
    arn_map = {entry.get("arn", ""): entry for entry in entries if entry.get("arn")}
    if not arn_map:
        return

    try:
        paginator = client.get_paginator("list_functions")
        for page in paginator.paginate():
            for function in page.get("Functions", []):
                arn = function.get("FunctionArn")
                entry = arn_map.get(arn)
                if not entry:
                    continue
                memory = function.get("MemorySize")
                timeout = function.get("Timeout")
                entry["details"].update(
                    {
                        "memory_mb": memory,
                        "timeout_seconds": timeout,
                        "runtime": function.get("Runtime"),
                    }
                )
                if memory is not None:
                    entry["size"] = f"{memory} MB"
    except (ClientError, BotoCoreError) as exc:
        sys.stderr.write(f"[WARN] {region} Lambda list_functions falhou: {exc}\n")


def enrich_sqs(session: boto3.Session, region: str, inventory: dict[str, list[dict]]):
    entries = inventory.get("sqs:queue")
    if not entries:
        return

    client = session.client("sqs", region_name=region)
    for entry in entries:
        arn = entry.get("arn", "")
        parts = arn.split(":")
        if len(parts) < 6:
            continue
        account_id = parts[4]
        queue_name = parts[5]
        queue_url = f"https://sqs.{region}.amazonaws.com/{account_id}/{queue_name}"
        try:
            response = client.get_queue_attributes(
                QueueUrl=queue_url,
                AttributeNames=[
                    "ApproximateNumberOfMessages",
                    "ApproximateNumberOfMessagesDelayed",
                    "MaximumMessageSize",
                    "MessageRetentionPeriod",
                ],
            )
        except (ClientError, BotoCoreError):
            continue
        attrs = response.get("Attributes", {})
        approx = attrs.get("ApproximateNumberOfMessages")
        max_size = attrs.get("MaximumMessageSize")
        retention = attrs.get("MessageRetentionPeriod")
        entry["details"].update(
            {
                "queue_url": queue_url,
                "approximate_messages": int(approx) if approx is not None else None,
                "maximum_message_size_bytes": int(max_size) if max_size is not None else None,
                "retention_seconds": int(retention) if retention is not None else None,
            }
        )
        if max_size:
            entry["size"] = f"{max_size} bytes por mensagem" if max_size else entry["size"]
        elif approx:
            entry["size"] = f"{approx} mensagens pendentes"


def enrich_sns(session: boto3.Session, region: str, inventory: dict[str, list[dict]]):
    entries = inventory.get("sns:topic")
    if not entries:
        return

    client = session.client("sns", region_name=region)
    for entry in entries:
        arn = entry.get("arn", "")
        if not arn:
            continue
        try:
            response = client.get_topic_attributes(TopicArn=arn)
        except (ClientError, BotoCoreError):
            continue
        attrs = response.get("Attributes", {})
        confirmed = attrs.get("SubscriptionsConfirmed")
        delivery = attrs.get("EffectiveDeliveryPolicy")
        entry["details"].update(
            {
                "subscriptions_confirmed": int(confirmed) if confirmed is not None else None,
                "effective_delivery_policy": delivery,
            }
        )
        if confirmed is not None:
            entry["size"] = f"{confirmed} assinaturas"


def write_region_inventory_json(output_dir: str, account_id: str, region: str, inventory: dict[str, list[dict]]):
    filename = f"report-{account_id}_{region}.json"
    output_path = os.path.join(output_dir, filename)
    with open(output_path, "w", encoding="utf-8") as handler:
        json.dump(inventory, handler, indent=2, sort_keys=True)
    return output_path


def autosize_columns(worksheet):
    for column_cells in worksheet.columns:
        cells = list(column_cells)
        if not cells:
            continue
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in cells)
        adjusted_width = min(max_length + 2, 80)
        worksheet.column_dimensions[get_column_letter(cells[0].column)].width = adjusted_width


def write_region_inventory_excel(output_dir: str, account_id: str, region: str, inventory: dict[str, list[dict]]):
    filename = f"report-{account_id}_{region}.xlsx"
    output_path = os.path.join(output_dir, filename)
    workbook = Workbook()

    ws_resources = workbook.active
    ws_resources.title = "Recursos"
    ws_resources.append(["resource_type", "family", "resource_name", "size", "arn", "tags"])

    for rtype, entries in sorted(inventory.items()):
        for entry in entries:
            ws_resources.append(
                [
                    rtype,
                    entry.get("family", ""),
                    entry.get("resource_name", ""),
                    entry.get("size", ""),
                    entry.get("arn", ""),
                    format_tags(entry.get("tags")),
                ]
            )
    ws_resources.freeze_panes = "A2"

    ws_summary = workbook.create_sheet("Resumo")
    ws_summary.append(["resource_type", "family", "count"])
    for rtype, entries in sorted(inventory.items()):
        ws_summary.append([rtype, rtype.split(":", 1)[0], len(entries)])
    ws_summary.freeze_panes = "A2"

    def write_service_sheet(prefix: str, title: str, type_label: str):
        service_entries = {
            rtype: entries
            for rtype, entries in inventory.items()
            if rtype.startswith(prefix)
        }
        if not service_entries:
            return
        worksheet = workbook.create_sheet(title)
        worksheet.append([type_label, "family", "resource_name", "size", "arn", "tags"])
        for subtype, entries in sorted(service_entries.items()):
            for entry in entries:
                worksheet.append(
                    [
                        subtype,
                        entry.get("family", ""),
                        entry.get("resource_name", ""),
                        entry.get("size", ""),
                        entry.get("arn", ""),
                        format_tags(entry.get("tags")),
                    ]
                )
        worksheet.append([])
        worksheet.append([type_label, "count"])
        for subtype, entries in sorted(service_entries.items()):
            worksheet.append([subtype, len(entries)])
        worksheet.freeze_panes = "A2"

    write_service_sheet("ec2:", "EC2", "ec2_type")
    write_service_sheet("rds:", "RDS", "rds_type")
    write_service_sheet("lambda:", "Lambda", "lambda_type")
    write_service_sheet("sqs:", "SQS", "sqs_type")
    write_service_sheet("sns:", "SNS", "sns_type")

    for worksheet in workbook.worksheets:
        autosize_columns(worksheet)

    workbook.save(output_path)
    return output_path


def iter_report_files(output_dir: Path):
    for path in sorted(output_dir.glob("report-*_*.json")):
        if not path.is_file():
            continue
        stem = path.stem
        if not stem.startswith("report-"):
            continue
        remainder = stem[len("report-") :]
        if "_" not in remainder:
            continue
        account, region = remainder.split("_", 1)
        yield path, account, region


def matches_service(service_filter: str, resource_type: str) -> bool:
    if ":" in service_filter:
        return resource_type == service_filter
    return resource_type.startswith(f"{service_filter}:")


def run_collect(args):
    page_size = max(1, min(args.page_size, 100))
    ensure_output_dir(args.output_dir)

    session_kwargs = {}
    if args.profile:
        session_kwargs["profile_name"] = args.profile
    session = boto3.Session(**session_kwargs)

    try:
        sts_client = session.client("sts")
        account_id = sts_client.get_caller_identity()["Account"]
    except (ClientError, BotoCoreError, KeyError) as exc:
        sys.stderr.write(f"Falha ao obter o ID da conta: {exc}\n")
        return 3

    try:
        regions = detect_regions(session, args.regions)
    except RuntimeError as exc:
        sys.stderr.write(f"{exc}\n")
        return 2

    success = False

    for region in regions:
        sys.stderr.write(f"Coletando recursos em {region}...\n")
        try:
            grouped, count, duplicates = collect_region_inventory(session, region, page_size)
        except (ClientError, BotoCoreError) as exc:
            sys.stderr.write(f"[WARN] Falha ao coletar {region}: {exc}\n")
            continue

        enrich_inventory(session, region, grouped)
        json_path = write_region_inventory_json(args.output_dir, account_id, region, grouped)
        excel_path = write_region_inventory_excel(args.output_dir, account_id, region, grouped)
        sys.stderr.write(
            f"{region}: {count} recursos únicos salvos em {json_path} e {excel_path}\n"
        )
        if duplicates:
            sys.stderr.write(
                f"[INFO] {region}: {duplicates} registros duplicados foram ignorados.\n"
            )
        success = True

    if not success:
        sys.stderr.write("Nenhum recurso coletado nas regiões solicitadas.\n")
        return 1

    return 0


def run_query(args):
    output_dir = Path(args.output_dir)
    if not output_dir.is_dir():
        sys.stderr.write(f"Diretório de inventário não encontrado: {output_dir}\n")
        return 2

    regions_filter = set(args.regions) if args.regions else None
    matches = 0

    for path, account, region in iter_report_files(output_dir):
        if regions_filter and region not in regions_filter:
            continue
        with path.open("r", encoding="utf-8") as handler:
            inventory = json.load(handler)
        for rtype, entries in inventory.items():
            if not matches_service(args.service, rtype):
                continue
            for entry in entries:
                record = {
                    "account": account,
                    "region": region,
                    "type": rtype,
                    "family": entry.get("family", ""),
                    "size": entry.get("size", ""),
                    "arn": entry.get("arn", ""),
                    "resource_name": entry.get("resource_name", ""),
                    "tags": entry.get("tags", {}),
                    "details": entry.get("details", {}),
                }
                print(json.dumps(record))
                matches += 1

    if matches == 0:
        sys.stderr.write(
            f"Nenhum recurso encontrado para o filtro \"{args.service}\".\n"
        )
        return 3

    return 0


def main():
    parser = build_parser()
    args = parser.parse_args()

    if args.command == "collect":
        return run_collect(args)
    if args.command == "query":
        return run_query(args)

    parser.error("Comando inválido.")


if __name__ == "__main__":  # pragma: no cover - CLI entry point
    sys.exit(main())
