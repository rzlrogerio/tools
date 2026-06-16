#!/usr/bin/env python3

"""Collect and query AWS resource inventory with JSON and Excel outputs."""

import argparse
import json
import os
import sys
from collections import defaultdict
from pathlib import Path

try:
    import boto3
    from botocore.exceptions import BotoCoreError, ClientError
except ImportError as exc:
    sys.stderr.write(f"boto3 is required to run this script: {exc}\n")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    sys.stderr.write(f"openpyxl is required to run this script: {exc}\n")
    sys.exit(1)


def build_parser():
    default_output = os.environ.get(
        "OUTPUT_DIR", os.path.expanduser("~/Download/inventario/aws")
    )
    parser = argparse.ArgumentParser(
        description=(
            "Coleta inventário de recursos AWS e permite consultas filtradas. "
            "Os dados são agrupados por tipo de recurso e exportados em JSON e Excel."
        )
    )

    subparsers = parser.add_subparsers(dest="command", required=True)

    collect_parser = subparsers.add_parser(
        "collect", help="Coleta o inventário e gera os relatórios."
    )
    collect_parser.add_argument(
        "--regions",
        nargs="*",
        help="Lista de regiões AWS. Padrão: todas as regiões suportadas pela API.",
    )
    collect_parser.add_argument(
        "--profile",
        help="Perfil de credenciais AWS a utilizar (opcional).",
    )
    collect_parser.add_argument(
        "--output-dir",
        default=default_output,
        help="Diretório de saída para os relatórios (padrão: ~/Download/inventario/aws).",
    )
    collect_parser.add_argument(
        "--page-size",
        type=int,
        default=100,
        help="Quantidade de recursos por requisição à API (máximo 100).",
    )

    query_parser = subparsers.add_parser(
        "query", help="Consulta relatórios existentes filtrando por serviço."
    )
    query_parser.add_argument(
        "service",
        help="Serviço (ex.: ec2) ou tipo completo (ex.: ec2:instance) a filtrar.",
    )
    query_parser.add_argument(
        "--regions",
        nargs="*",
        help="Filtra resultados por região (opcional).",
    )
    query_parser.add_argument(
        "--output-dir",
        default=default_output,
        help="Diretório contendo os relatórios gerados (padrão: ~/Download/inventario/aws).",
    )

    return parser


def ensure_output_dir(path):
    os.makedirs(path, exist_ok=True)


def detect_regions(session, cli_regions):
    if cli_regions:
        return cli_regions
    available = session.get_available_regions("resourcegroupstaggingapi")
    if not available:
        raise RuntimeError("Nenhuma região disponível para resourcegroupstaggingapi.")
    return available


def resource_type_from_arn(arn):
    if not arn or not arn.startswith("arn:"):
        return "unknown"
    parts = arn.split(":", 5)
    if len(parts) < 6:
        return "unknown"
    service = parts[2] or "unknown"
    resource = parts[5]
    primary = ""
    if service == "s3" and resource and "/" not in resource and ":" not in resource:
        primary = "bucket"
    else:
        for separator in ("/", ":"):
            if separator in resource:
                primary = resource.split(separator, 1)[0]
                break
        if not primary:
            primary = resource if resource else "resource"
    return f"{service}:{primary}"


def collect_region_inventory(session, region, page_size):
    client = session.client("resourcegroupstaggingapi", region_name=region)
    grouped = defaultdict(dict)  # resource_type -> {arn: entry}
    duplicates = 0
    paginator = client.get_paginator("get_resources")

    for page in paginator.paginate(ResourcesPerPage=page_size):
        for mapping in page.get("ResourceTagMappingList", []):
            arn = mapping.get("ResourceARN")
            if not arn:
                continue
            resource_type = resource_type_from_arn(arn)
            tags = {}
            for tag in mapping.get("Tags", []):
                key = tag.get("Key")
                if key is None:
                    continue
                tags[key] = tag.get("Value", "")

            existing = grouped[resource_type].get(arn)
            if existing:
                duplicates += 1
                if tags:
                    existing_tags = existing.get("tags") or {}
                    existing_tags.update(tags)
                    existing["tags"] = existing_tags
                continue

            grouped[resource_type][arn] = {"arn": arn, "tags": tags}

    grouped_lists = {
        rtype: list(entries.values()) for rtype, entries in grouped.items()
    }
    total_unique = sum(len(entries) for entries in grouped_lists.values())
    return grouped_lists, total_unique, duplicates


def resource_display_name(arn, tags):
    if tags:
        name = tags.get("Name") or tags.get("name")
        if name:
            return name
    if not arn:
        return ""
    resource_id = arn.rsplit("/", 1)[-1]
    resource_id = resource_id.rsplit(":", 1)[-1]
    return resource_id


def format_tags(tags):
    if not tags:
        return ""
    return ";".join(
        f"{key}={'' if value is None else value}"
        for key, value in sorted(tags.items())
    )


def write_region_inventory_json(output_dir, account_id, region, inventory):
    filename = f"report-{account_id}_{region}.json"
    output_path = os.path.join(output_dir, filename)
    with open(output_path, "w", encoding="utf-8") as handler:
        json.dump(inventory, handler, indent=2, sort_keys=True)
    return output_path


def autosize_columns(worksheet):
    for column_cells in worksheet.columns:
        cells = list(column_cells)
        if not cells:
            continue
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in cells)
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[get_column_letter(cells[0].column)].width = adjusted_width


def write_region_inventory_excel(output_dir, account_id, region, inventory):
    filename = f"report-{account_id}_{region}.xlsx"
    output_path = os.path.join(output_dir, filename)
    workbook = Workbook()

    ws_resources = workbook.active
    ws_resources.title = "Recursos"
    ws_resources.append(["resource_type", "resource_name", "arn", "tags"])

    for rtype, entries in sorted(inventory.items()):
        for entry in entries:
            arn = entry.get("arn", "")
            tags = entry.get("tags", {}) or {}
            ws_resources.append(
                [rtype, resource_display_name(arn, tags), arn, format_tags(tags)]
            )
    ws_resources.freeze_panes = "A2"

    ws_summary = workbook.create_sheet("Resumo")
    ws_summary.append(["resource_type", "count"])
    for rtype, entries in sorted(inventory.items()):
        ws_summary.append([rtype, len(entries)])
    ws_summary.freeze_panes = "A2"

    def write_service_sheet(service_prefix, sheet_title, type_label):
        service_entries = {
            rtype: entries
            for rtype, entries in inventory.items()
            if rtype.startswith(service_prefix)
        }
        if not service_entries:
            return

        worksheet = workbook.create_sheet(sheet_title)
        worksheet.append([type_label, "resource_name", "arn", "tags"])
        for subtype, entries in sorted(service_entries.items()):
            for entry in entries:
                arn = entry.get("arn", "")
                tags = entry.get("tags", {}) or {}
                worksheet.append(
                    [subtype, resource_display_name(arn, tags), arn, format_tags(tags)]
                )
        worksheet.append([])
        worksheet.append([type_label, "count"])
        for subtype, entries in sorted(service_entries.items()):
            worksheet.append([subtype, len(entries)])
        worksheet.freeze_panes = "A2"

    write_service_sheet("ec2:", "EC2", "ec2_type")
    write_service_sheet("rds:", "RDS", "rds_type")
    write_service_sheet("lambda:", "Lambda", "lambda_type")
    write_service_sheet("sqs:", "SQS", "sqs_type")
    write_service_sheet("sns:", "SNS", "sns_type")

    for worksheet in workbook.worksheets:
        autosize_columns(worksheet)

    workbook.save(output_path)
    return output_path


def run_collect(args):
    page_size = max(1, min(args.page_size, 100))
    ensure_output_dir(args.output_dir)

    session_kwargs = {}
    if args.profile:
        session_kwargs["profile_name"] = args.profile
    session = boto3.Session(**session_kwargs)

    try:
        sts_client = session.client("sts")
        account_id = sts_client.get_caller_identity()["Account"]
    except (ClientError, BotoCoreError, KeyError) as exc:
        sys.stderr.write(f"Falha ao obter o ID da conta: {exc}\n")
        return 3

    try:
        regions = detect_regions(session, args.regions)
    except RuntimeError as exc:
        sys.stderr.write(f"{exc}\n")
        return 2

    success = False

    for region in regions:
        sys.stderr.write(f"Coletando recursos em {region}...\n")
        try:
            grouped, count, duplicates = collect_region_inventory(
                session, region, page_size
            )
        except (ClientError, BotoCoreError) as exc:
            sys.stderr.write(f"[WARN] Falha ao coletar {region}: {exc}\n")
            continue

        ordered = {rtype: entries for rtype, entries in sorted(grouped.items())}
        json_path = write_region_inventory_json(args.output_dir, account_id, region, ordered)
        excel_path = write_region_inventory_excel(args.output_dir, account_id, region, ordered)
        sys.stderr.write(
            f"{region}: {count} recursos únicos salvos em {json_path} e {excel_path}\n"
        )
        if duplicates:
            sys.stderr.write(
                f"[INFO] {region}: {duplicates} registros duplicados foram ignorados.\n"
            )
        success = True

    if not success:
        sys.stderr.write("Nenhum recurso coletado nas regiões solicitadas.\n")
        return 1

    return 0


def iter_report_files(output_dir):
    base_path = Path(output_dir)
    for path in sorted(base_path.glob("report-*_*.json")):
        if not path.is_file():
            continue
        stem = path.stem  # report-<account>_<region>
        if not stem.startswith("report-"):
            continue
        remainder = stem[len("report-") :]
        if "_" not in remainder:
            continue
        account, region = remainder.split("_", 1)
        yield path, account, region


def matches_service(service_filter, resource_type):
    if ":" in service_filter:
        return resource_type == service_filter
    return resource_type.startswith(f"{service_filter}:")


def run_query(args):
    output_dir = Path(args.output_dir)
    if not output_dir.is_dir():
        sys.stderr.write(f"Diretório de inventário não encontrado: {output_dir}\n")
        return 2

    regions_filter = set(args.regions) if args.regions else None
    matches = 0

    for path, account, region in iter_report_files(output_dir):
        if regions_filter and region not in regions_filter:
            continue
        with path.open("r", encoding="utf-8") as handler:
            inventory = json.load(handler)
        for rtype, entries in inventory.items():
            if not matches_service(args.service, rtype):
                continue
            for entry in entries:
                arn = entry.get("arn", "")
                tags = entry.get("tags", {}) or {}
                record = {
                    "account": account,
                    "region": region,
                    "type": rtype,
                    "arn": arn,
                    "resource_name": resource_display_name(arn, tags),
                    "tags": tags,
                }
                print(json.dumps(record))
                matches += 1

    if matches == 0:
        sys.stderr.write(
            f"Nenhum recurso encontrado para o filtro \"{args.service}\".\n"
        )
        return 3

    return 0


def main():
    parser = build_parser()
    args = parser.parse_args()

    if args.command == "collect":
        return run_collect(args)
    if args.command == "query":
        return run_query(args)

    parser.error("Comando inválido.")


if __name__ == "__main__":  # pragma: no cover - CLI entry point
    sys.exit(main())
